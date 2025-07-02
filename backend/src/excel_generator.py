import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from typing import Dict, List
import logging
from datetime import datetime

from .models import ProcessingResult, DormantCustomer, SalespersonSummary

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Advanced Excel report generator with formatting and charts."""
    
    def __init__(self):
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.title_font = Font(bold=True, size=14, color="2F5597")
        self.subtitle_font = Font(bold=True, size=11)
        self.highlight_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_report(self, result: ProcessingResult, output_path: str) -> str:
        """Generate comprehensive Excel report."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, result)
        self._create_salesperson_sheets(wb, result)
        self._create_consolidated_sheet(wb, result)
        self._create_analytics_dashboard(wb, result)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report generated: {output_path}")
        
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, result: ProcessingResult):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "Well Crafted Wine & Beverage Co."
        ws['A1'].font = Font(bold=True, size=16, color="8D4004")
        ws['A2'] = "Dormant Customer Analysis Report"
        ws['A2'].font = Font(bold=True, size=14, color="2F5597")
        ws['A3'] = f"Generated: {result.processing_timestamp.strftime('%B %d, %Y at %I:%M %p')}"
        ws['A3'].font = Font(size=10, italic=True)
        
        # Key Metrics
        ws['A5'] = "📊 KEY METRICS"
        ws['A5'].font = self.title_font
        
        metrics = [
            ("Total Dormant Customers", len(result.dormant_customers)),
            ("Total Value at Risk", f"${result.summary['total_value_at_risk']:,.2f}"),
            ("Average Churn Risk", f"{result.summary['average_churn_risk']:.1%}"),
            ("Data Accuracy Score", f"{result.data_accuracy_score:.1%}"),
            ("Customers Analyzed", result.total_customers_analyzed)
        ]
        
        for i, (metric, value) in enumerate(metrics, 6):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = self.subtitle_font
        
        # AI Strategic Insights
        ws['A12'] = "🎯 AI-POWERED STRATEGIC INSIGHTS"
        ws['A12'].font = self.title_font
        
        row = 13
        for key, insight in result.insights.items():
            ws[f'A{row}'] = f"• {insight}"
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1
        
        # Salesperson Summary Table
        ws[f'A{row + 1}'] = "👥 SALESPERSON PERFORMANCE SUMMARY"
        ws[f'A{row + 1}'].font = self.title_font
        
        # Headers
        headers = ["Salesperson", "Dormant Customers", "Value at Risk", "High Value", "Quick Wins", "Avg Churn Risk"]
        header_row = row + 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data rows
        for i, summary in enumerate(result.salesperson_summaries, header_row + 1):
            data = [
                summary.salesperson,
                summary.dormant_customer_count,
                f"${summary.total_value_at_risk:,.2f}",
                summary.high_value_dormant_count,
                summary.quick_win_count,
                f"{summary.average_churn_risk:.1%}"
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.border
                
                # Highlight high-risk reps
                if summary.average_churn_risk > 0.7:
                    cell.fill = self.warning_fill
                elif summary.total_value_at_risk > 5000:
                    cell.fill = self.highlight_fill
        
        # Data Quality Report
        quality_row = header_row + len(result.salesperson_summaries) + 3
        ws[f'A{quality_row}'] = "📋 DATA QUALITY REPORT"
        ws[f'A{quality_row}'].font = self.title_font
        
        quality_metrics = [
            ("Total Records Processed", result.data_quality_report["total_records"]),
            ("Valid Records", result.data_quality_report["valid_records"]),
            ("Data Completeness", f"{result.data_quality_report['data_completeness_score']:.1%}"),
            ("Missing Customer Mappings", result.data_quality_report["missing_customer_mappings"])
        ]
        
        for i, (metric, value) in enumerate(quality_metrics, quality_row + 1):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = self.subtitle_font
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_salesperson_sheets(self, wb: Workbook, result: ProcessingResult):
        """Create individual sheets for each salesperson."""
        # Group customers by salesperson
        rep_customers = {}
        for customer in result.dormant_customers:
            rep = customer.salesperson
            if rep not in rep_customers:
                rep_customers[rep] = []
            rep_customers[rep].append(customer)
        
        for rep, customers in rep_customers.items():
            # Clean sheet name
            sheet_name = self._clean_sheet_name(f"{rep}_Dormant")
            ws = wb.create_sheet(sheet_name)
            
            # Title
            ws['A1'] = f"Dormant Customers - {rep}"
            ws['A1'].font = Font(bold=True, size=14, color="2F5597")
            
            # Summary for this rep
            summary = next((s for s in result.salesperson_summaries if s.salesperson == rep), None)
            if summary:
                ws['A3'] = f"Total Dormant Customers: {summary.dormant_customer_count}"
                ws['A4'] = f"Total Value at Risk: ${summary.total_value_at_risk:,.2f}"
                ws['A5'] = f"Average Churn Risk: {summary.average_churn_risk:.1%}"
            
            # Table headers
            headers = [
                "Customer", "Last Order Date", "Days Since Order", "6-Month Value",
                "Order Count", "Avg Order Value", "Churn Risk", "CLV", "Preferred Products"
            ]
            
            header_row = 7
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            # Customer data
            for i, customer in enumerate(sorted(customers, key=lambda x: x.churn_risk_score, reverse=True), header_row + 1):
                data = [
                    customer.customer,
                    customer.last_order_date.strftime('%m/%d/%Y'),
                    customer.days_since_order,
                    f"${customer.total_6_month_value:,.2f}",
                    customer.order_count_6_months,
                    f"${customer.average_order_value:,.2f}",
                    f"{customer.churn_risk_score:.1%}",
                    f"${customer.customer_lifetime_value:,.2f}",
                    ", ".join(customer.preferred_products[:3])
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=i, column=col, value=value)
                    cell.border = self.border
                    
                    # Color coding by risk
                    if customer.churn_risk_score > 0.8:
                        cell.fill = self.warning_fill
                    elif customer.churn_risk_score > 0.6:
                        cell.fill = self.highlight_fill
            
            # Auto-fit columns
            self._auto_fit_columns(ws)
    
    def _create_consolidated_sheet(self, wb: Workbook, result: ProcessingResult):
        """Create consolidated view of all dormant customers."""
        ws = wb.create_sheet("All Dormant Customers")
        
        # Title
        ws['A1'] = "All Dormant Customers - Consolidated View"
        ws['A1'].font = Font(bold=True, size=14, color="2F5597")
        
        # Headers
        headers = [
            "Customer", "Salesperson", "Last Order Date", "Days Since Order",
            "6-Month Value", "Churn Risk", "Customer LTV", "Priority"
        ]
        
        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Sort customers by value at risk
        sorted_customers = sorted(result.dormant_customers, key=lambda x: x.total_6_month_value, reverse=True)
        
        for i, customer in enumerate(sorted_customers, header_row + 1):
            # Determine priority
            if customer.churn_risk_score > 0.8 or customer.total_6_month_value > 2000:
                priority = "HIGH"
            elif customer.churn_risk_score > 0.6 or customer.total_6_month_value > 1000:
                priority = "MEDIUM"
            else:
                priority = "LOW"
            
            data = [
                customer.customer,
                customer.salesperson,
                customer.last_order_date.strftime('%m/%d/%Y'),
                customer.days_since_order,
                f"${customer.total_6_month_value:,.2f}",
                f"{customer.churn_risk_score:.1%}",
                f"${customer.customer_lifetime_value:,.2f}",
                priority
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = self.border
                
                # Color coding
                if priority == "HIGH":
                    cell.fill = self.warning_fill
                elif priority == "MEDIUM":
                    cell.fill = self.highlight_fill
        
        self._auto_fit_columns(ws)
    
    def _create_analytics_dashboard(self, wb: Workbook, result: ProcessingResult):
        """Create analytics dashboard with charts."""
        ws = wb.create_sheet("Analytics Dashboard")
        
        # Title
        ws['A1'] = "Analytics Dashboard"
        ws['A1'].font = Font(bold=True, size=16, color="2F5597")
        
        # Risk distribution
        ws['A3'] = "Churn Risk Distribution"
        ws['A3'].font = self.title_font
        
        # Calculate risk buckets
        high_risk = len([c for c in result.dormant_customers if c.churn_risk_score > 0.7])
        medium_risk = len([c for c in result.dormant_customers if 0.4 <= c.churn_risk_score <= 0.7])
        low_risk = len([c for c in result.dormant_customers if c.churn_risk_score < 0.4])
        
        risk_data = [
            ["Risk Level", "Customer Count"],
            ["High Risk (>70%)", high_risk],
            ["Medium Risk (40-70%)", medium_risk],
            ["Low Risk (<40%)", low_risk]
        ]
        
        for i, row in enumerate(risk_data, 4):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 4:  # Header
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                cell.border = self.border
        
        # Value distribution
        ws['A10'] = "Value at Risk by Salesperson"
        ws['A10'].font = self.title_font
        
        value_data = [["Salesperson", "Value at Risk"]]
        for summary in result.salesperson_summaries[:10]:  # Top 10
            value_data.append([summary.salesperson, float(summary.total_value_at_risk)])
        
        for i, row in enumerate(value_data, 11):
            for j, value in enumerate(row, 1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 11:  # Header
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                if j == 2 and i > 11:  # Format currency
                    cell.value = f"${value:,.2f}"
                cell.border = self.border
        
        self._auto_fit_columns(ws)
    
    def _clean_sheet_name(self, name: str) -> str:
        """Clean sheet name to meet Excel requirements."""
        # Remove invalid characters
        invalid_chars = ['[', ']', '*', '?', ':', '/', '\\']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Limit to 31 characters
        return name[:31]
    
    def _auto_fit_columns(self, ws):
        """Auto-fit column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width